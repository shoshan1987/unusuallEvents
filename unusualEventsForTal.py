import openpyxl
import datetime
import time
from openpyxl.styles import NamedStyle
from copy import copy
from openpyxl.utils import FORMULAE
import datetime
from datetime import timedelta

#now = datetime.datetime.now()
today = datetime.datetime.today()


d1 = today.strftime("%d/%m/%Y")
file = input("What is the file name you want to me  ?\n")
file = openpyxl.load_workbook(file + '.xlsx')

file.active


print("I activated file")

file.create_sheet('dataAfterManipulation')
file.save(r"C:\Users\shoshana\PycharmProjects\pandas\UnusualEvents\EventReportFormatAugust_afterManipulation3.xlsx")
wb = openpyxl.load_workbook(r"C:\Users\shoshana\PycharmProjects\pandas\styling.xlsx")
origin = file['row data'].cell
destination = file['dataAfterManipulation'].cell


print("I defined vars")


def setHourASHour():

    date_style = NamedStyle(name='datetime', number_format='HH:MM:SS')
    for i in range(7,257):
        file['row data'].cell(row=i, column=8).style = date_style
        file['row data'].cell(row=i, column=8).number_format = copy(wb['Sheet']['B1'].number_format)
        file['row data'].cell(row=i, column=9).style = date_style
        file['row data'].cell(row=i, column=9).number_format = copy(wb['Sheet']['B1'].number_format)

    file.save((r"C:\Users\shoshana\PycharmProjects\pandas\UnusualEvents\EventReportFormatAugust_afterManipulation3.xlsx"))
    return


def setLables():
    print("started setting Lables")
    fromWhereToCopy=[10, 8, 9, 13]
    whereToPaste = [4, 2 , 3 ,1]

    for i in range(0, len(fromWhereToCopy)):
        file['dataAfterManipulation'].cell(row=1, column=whereToPaste[i]).value = copy(file['row data'].cell(row=6, column=fromWhereToCopy[i]).value)
        file['dataAfterManipulation'].cell(row=1, column=whereToPaste[i]).font = copy(file['row data'].cell(row=6, column=fromWhereToCopy[i]).font)
        file['dataAfterManipulation'].cell(row=1, column=whereToPaste[i]).border = copy(file['row data'].cell(row=6, column=fromWhereToCopy[i]).border)
        file['dataAfterManipulation'].cell(row=1, column=whereToPaste[i]).fill = copy(file['row data'].cell(row=6, column=fromWhereToCopy[i]).fill)
        file['dataAfterManipulation'].cell(row=1, column=whereToPaste[i]).number_format = copy(file['row data'].cell(row=6, column=fromWhereToCopy[i]).number_format)
        file['dataAfterManipulation'].cell(row=1, column=whereToPaste[i]).protection = copy(file['row data'].cell(row=6, column=fromWhereToCopy[i]).protection)
        file['dataAfterManipulation'].cell(row=1, column=whereToPaste[i]).alignment = copy(file['row data'].cell(row=6, column=fromWhereToCopy[i]).alignment)

    file.save((r"C:\Users\shoshana\PycharmProjects\pandas\UnusualEvents\EventReportFormatAugust_afterManipulation3.xlsx"))
    return

def copyPasteRelevantDataAndSortIt():
    print("started copyPasteRelevantDataAndSortIt")
    # file = openpyxl.load_workbook(r"C:\Users\shoshana\PycharmProjects\pandas\UnusualEvents\EventReportFormatAugust.xlsx", data_only= True)
    eventType =[]
    startStopTime =[]
    endStopTime =[]
    NoneArr = [None] * 100

    for i in range(7,250):
        realValueColumn8 =  (file['row data'].cell(row=i, column=8).value)
        realValueColumn9 = (file['row data'].cell(row=i, column=9).value)
        realValueColumn10 = (file['row data'].cell(row=i, column=10).value)

        if type(file['row data'].cell(row=i, column=8).value) is not str and (file['row data'].cell(row=i, column=8).value) is not None:
            column8 = (file['row data'].cell(row=i, column=8).value).strftime("%d/%m/%Y")
        else:
            column8 = (file['row data'].cell(row=i, column=8).value)


        if type(file['row data'].cell(row=i, column=9).value) is not str and (file['row data'].cell(row=i, column=9).value) is not None:
            column9 = (file['row data'].cell(row=i, column=9).value).strftime("%d/%m/%Y")
        else:
            column9 = (file['row data'].cell(row=i, column=9).value)

        if type(file['row data'].cell(row=i, column=10).value) is not str and (file['row data'].cell(row=i, column=10).value) is not None:
            column10 = (file['row data'].cell(row=i, column=10).value).strftime("%d/%m/%Y")
        else:
            column10 = (file['row data'].cell(row=i, column=10).value)



        if file['row data'].cell(row=i, column=10).value is not None:
            if "אין"not in column8 and "אין" not in column9  and "לא היו אירועים חריגים" not in file['row data'].cell(row=i, column=4).value:

                if "דילוג" not in column8 and "עציר" not in column8:
                    if column10 is not None:
                        eventType.append(realValueColumn10)
                        startStopTime.append(realValueColumn8)
                        endStopTime.append(realValueColumn9)
                        # print((file['row data'].cell(row=i, column = 10).value))
                        # print((file['row data'].cell(row=i, column = 8).value))
                        # print((file['row data'].cell(row=i, column = 9).value))
                        # print("&&&&&&&&&&&&&&&&&&&")
                        # print(i)
                        # time.sleep(3)


                    elif "דילוג"in column8 and "עציר" in column8:
                        if len (file['row data'].cell(row=i, column=8).value) == 23:
                           strForConvertion = column8
                           strForConvertion = strForConvertion[18:22]
                           realValueColumn8 = strForConvertion
                           eventType.append(realValueColumn10)
                           startStopTime.append(realValueColumn8)
                           endStopTime.append(realValueColumn9)

                    elif "דילוג" not in column8 and "עציר" in column8:
                        strForConvertion = column8
                        for a in range(0, 6):
                            a = str(a)
                            for b in range(0, 10):
                                b = str(b)
                                for c in range(0, 6):
                                    c = str(c)
                                    for d in range(0, 10):
                                        d = str(d)

                                        if (a + b + ":" +c + d) in strForConvertion:
                                            eventType.append(realValueColumn10)
                                            startStopTime.append(a + b + ":" + c + d)
                                            endStopTime.append(realValueColumn9)

                    else:
                        print("huston! we have a problem")


    print("finishToCopyToArrs")

    eventType, startStopTime = sortingArraysByMainArray(eventType, startStopTime, doYouWantToSortMainArr=1)
    eventType, endStopTime = sortingArraysByMainArray(eventType, endStopTime, doYouWantToSortMainArr=0)
    print("finish to sort")

    eventTypeWithNone = eventType + NoneArr
    startStopTimeWithNone = startStopTime + NoneArr
    endStopTimeWothNone = endStopTime + NoneArr


    eventType = cleanNeedlessNoneFromArr(eventTypeWithNone)
    startStopTime = cleanNeedlessNoneFromArr(startStopTimeWithNone)
    endStopTime = cleanNeedlessNoneFromArr(endStopTimeWothNone)

    print("finish cleanNeedlessNoneFromArr")


    for i in range(2, len(eventType)+2):
        if eventType[i-2] is not None:
            file['dataAfterManipulation'].cell(row=i, column=4).value = eventType[i-2]

        if startStopTime[i - 2] is not None:
            file['dataAfterManipulation'].cell(row=i, column=2).value = startStopTime[i-2]

        if endStopTime[i - 2] is not None:
            file['dataAfterManipulation'].cell(row=i, column=3).value = endStopTime[i-2]

    print("finish pasting")
    file.save(r"C:\Users\shoshana\PycharmProjects\pandas\UnusualEvents\EventReportFormatAugust_afterManipulation3.xlsx")
    print("finish saving first time")
    return

def sortingArraysByMainArray(mainArr, secondaryArray, doYouWantToSortMainArr):
   twoArrssorted = sorted(zip(mainArr, secondaryArray))
   secondaryArray = [secondaryArray for mainArr, secondaryArray in twoArrssorted]
   if doYouWantToSortMainArr == 0:
       mainArr = sorted(mainArr)

   return(mainArr ,secondaryArray)

def cleanNeedlessNoneFromArr(Arr):
    for i in range(0, len(Arr)):
        if Arr[i] is not None:
            theBiggestINdexThatIsNotEmpty = i
    Arr = Arr[0: theBiggestINdexThatIsNotEmpty + 1]


    return(Arr)

def addFormula():
    print("started addind formoula")
    for m in range(2,250):
        m = str(m)
        if (file['dataAfterManipulation'].cell(row= int(m), column =4).value) is not None:
            date_style = NamedStyle(name='datetime', number_format='HH:MM:SS')

            file['dataAfterManipulation'].cell(row=int(m), column=1).number_format = copy(wb['Sheet']['B1'].number_format)

            startStopTime = "B" + m
            endStopTime = "C" + m
            formula = "=" + endStopTime + "-" + startStopTime
            #+ "= '#VALUE!', 0 ," + startStopTime + "-" + endStopTime + ")"
            m = int(m)
            file['dataAfterManipulation'].cell(row=m , column= 1).value = formula
    file.save(r"C:\Users\shoshana\PycharmProjects\pandas\UnusualEvents\EventReportFormatAugust_afterManipulation3.xlsx")

    print("finish adding formoula")
    return

def finalManipulation():
    print("started finalManipulation")
    eventTypeUnique=[]
    startStopUniqueUnique=copy(wb['Sheet']['C1'].value)
    endStopUniqueUnique=copy(wb['Sheet']['C1'].value)
    totalTimeCounter = 0

    totalTimeArr=[]
    numberOfNones = 0
    for i in range (2,250):
        if file['dataAfterManipulation'].cell(row=i, column=4).value is not None:
            eventTypeForAppending = file['dataAfterManipulation'].cell(row=i, column=4).value
            if i > 2 and i < 250:
                currentIndex = ((file['dataAfterManipulation'].cell(row=i, column=4).value))
                previousIndex = ((file['dataAfterManipulation'].cell(row=i - 1, column=4).value))
                nextIndex = ((file['dataAfterManipulation'].cell(row=i + 1, column=4).value))
                previousPreviousIndex = ((file['dataAfterManipulation'].cell(row=i - 2, column=4).value))
                if " " in previousPreviousIndex:
                    previousPreviousIndex = previousPreviousIndex.replace(" ", "")

                if " " in currentIndex:
                    currentIndex = currentIndex.replace(" ", "")
                if " " in previousIndex:
                    previousIndex = previousIndex.replace(" ", "")
                if currentIndex == previousIndex:
                    eventTypeUnique.append(None)

                    if type(file['dataAfterManipulation'].cell(row=i, column=3).value) is not str:
                        strForConvertion = file['dataAfterManipulation'].cell(row=i, column=3).value.strftime("%d/%m/%Y %H:%M:%S")
                        if '01/01/1900' in strForConvertion:
                            strForConvertion= strForConvertion.replace("01/01/1900","")
                    else:
                        strForConvertion = file['dataAfterManipulation'].cell(row=i, column=3).value

                    endTimeForCalculation = datetime.datetime.strptime(d1 + " " + strForConvertion, "%d/%m/%Y %H:%M:%S")
                    if type(file['dataAfterManipulation'].cell(row=i, column=2).value) is not str:
                        strForConvertion = file['dataAfterManipulation'].cell(row=i, column=2).value.strftime("%d/%m/%Y %H:%M:%S")
                        if '01/01/1900' in strForConvertion:
                            strForConvertion= strForConvertion.replace("01/01/1900","")
                    else:
                        strForConvertion = file['dataAfterManipulation'].cell(row=i, column=2).value
                    stratTimeForCalculation = datetime.datetime.strptime(d1 + " " + strForConvertion, "%d/%m/%Y %H:%M:%S")

                    totalTime = totalTime + endTimeForCalculation - stratTimeForCalculation


                else:
                    eventTypeUnique.append(file['dataAfterManipulation'].cell(row=i-1, column=4).value)

                    #if previousPreviousIndex == currentIndex:
                    if i > 3:
                        print(totalTime)
                        print(currentIndex)
                        print(len(currentIndex))
                        print(previousIndex)
                        print(len(previousIndex))

                        totalTimeArr.append(totalTime)
                    if type(file['dataAfterManipulation'].cell(row=i, column=3).value) is not str:
                        strForConvertion = file['dataAfterManipulation'].cell(row=i, column=3).value.strftime("%d/%m/%Y %H:%M:%S")
                        if '01/01/1900' in strForConvertion:
                            strForConvertion = strForConvertion.replace("01/01/1900", "")
                    else:
                        strForConvertion = file['dataAfterManipulation'].cell(row=i, column=3).value
                    endTimeForCalculation = datetime.datetime.strptime(d1 + " " + strForConvertion, "%d/%m/%Y %H:%M:%S")
                    if type(file['dataAfterManipulation'].cell(row=i, column=2).value) is not  str:
                        strForConvertion = file['dataAfterManipulation'].cell(row=i, column=2).value.strftime("%d/%m/%Y %H:%M:%S")
                        if '01/01/1900' in strForConvertion:
                            strForConvertion = strForConvertion.replace("01/01/1900", "")
                    else:
                        strForConvertion = file['dataAfterManipulation'].cell(row=i, column=2).value
                    stratTimeForCalculation = datetime.datetime.strptime(d1 + " " + strForConvertion, "%d/%m/%Y %H:%M:%S")

                    totalTime = endTimeForCalculation - stratTimeForCalculation


            elif i ==2:
                if type(file['dataAfterManipulation'].cell(row=i, column=3).value) is not str:

                    strForConvertion = file['dataAfterManipulation'].cell(row=i, column=3).value.strftime("%d/%m/%Y %H:%M:%S")

                    if '01/01/1900' in strForConvertion:
                        strForConvertion = strForConvertion.replace("01/01/1900", "")
                else:
                    strForConvertion = file['dataAfterManipulation'].cell(row=i, column=3).value
                    print(file['dataAfterManipulation'].cell(row=i, column=3).value)
                endTimeForCalculation = datetime.datetime.strptime(d1 + " " + strForConvertion, "%d/%m/%Y %H:%M:%S")
                if type(file['dataAfterManipulation'].cell(row=i, column=2).value) is not str:
                    strForConvertion = file['dataAfterManipulation'].cell(row=i, column=2).value.strftime("%d/%m/%Y %H:%M:%S")
                    if '01/01/1900' in strForConvertion:
                        strForConvertion = strForConvertion.replace("01/01/1900", "")
                else:
                    strForConvertion = file['dataAfterManipulation'].cell(row=i, column=2).value
                stratTimeForCalculation = datetime.datetime.strptime(d1 + " " + strForConvertion, "%d/%m/%Y %H:%M:%S")
                totalTime = endTimeForCalculation - stratTimeForCalculation

        else:
            numberOfNones = numberOfNones +1
            if numberOfNones== 10:
                totalTimeArr.append(totalTime)
                eventTypeUnique.append(eventTypeForAppending)
    file.create_sheet('dataAfterManipulationFinal')
    file.save(r"C:\Users\shoshana\PycharmProjects\pandas\UnusualEvents\EventReportFormatAugust_afterManipulation3.xlsx")

    for i in range (2, len(eventTypeUnique)+2):
        if eventTypeUnique[i-2] is not None:
            print(eventTypeUnique[i-2])
            print(len(eventTypeUnique[i-2]))
            (file['dataAfterManipulationFinal'].cell(row=i, column=2).value) = eventTypeUnique[i-2]

    for i in range(2, len(totalTimeArr)+2):
        if totalTimeArr[i-2] is not None:
            (file['dataAfterManipulationFinal'].cell(row=i+1, column=1).value) = totalTimeArr[i - 2]

    file['dataAfterManipulationFinal'].cell(row=1, column=2).value = copy(file['row data'].cell(row=6, column=10).value)
    file['dataAfterManipulationFinal'].cell(row=1, column=2).font = copy(file['row data'].cell(row=6, column=10).font)
    file['dataAfterManipulationFinal'].cell(row=1, column=2).border = copy(file['row data'].cell(row=6, column=10).border)
    file['dataAfterManipulationFinal'].cell(row=1, column=2).fill = copy(file['row data'].cell(row=6, column=10).fill)
    file['dataAfterManipulationFinal'].cell(row=1, column=2).number_format = copy(file['row data'].cell(row=6, column=10).number_format)
    file['dataAfterManipulationFinal'].cell(row=1, column=2).protection = copy(file['row data'].cell(row=6, column=10).protection)
    file['dataAfterManipulationFinal'].cell(row=1, column=2).alignment = copy(file['row data'].cell(row=6, column=10).alignment)

    file['dataAfterManipulationFinal'].cell(row=1, column=1).value = copy(file['row data'].cell(row=6, column=13).value)
    file['dataAfterManipulationFinal'].cell(row=1, column=1).font = copy(file['row data'].cell(row=6, column=13).font)
    file['dataAfterManipulationFinal'].cell(row=1, column=1).border = copy(file['row data'].cell(row=6, column=13).border)
    file['dataAfterManipulationFinal'].cell(row=1, column=1).fill = copy(file['row data'].cell(row=6, column=13).fill)
    file['dataAfterManipulationFinal'].cell(row=1, column=1).number_format = copy(file['row data'].cell(row=6, column=13).number_format)
    file['dataAfterManipulationFinal'].cell(row=1, column=1).protection = copy(file['row data'].cell(row=6, column=13).protection)
    file['dataAfterManipulationFinal'].cell(row=1, column=1).alignment = copy(file['row data'].cell(row=6, column=13).alignment)

    print("finished final manipulation")
    return



def main():
    setHourASHour
    setLables()
    copyPasteRelevantDataAndSortIt()
    addFormula()
    finalManipulation()
    print("started final saving")
    file.save(r"C:\Users\shoshana\PycharmProjects\pandas\UnusualEvents\EventReportFormatAugust_afterManipulation3.xlsx")
    print("finished final saving")
    return


main()

